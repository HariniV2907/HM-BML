from kivy.app import App
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.gridlayout import GridLayout
from openpyxl import load_workbook
import os

# File paths
USER_DATA_FILE = 'C:/Users/Sakth/Downloads/Users.xlsx'
BILL_DATA_FILE = 'C:/Users/Sakth/Downloads/bills.xlsx'  # Use raw string literal for Windows path

# Initialize Excel file if not exists
if not os.path.exists(USER_DATA_FILE):
    wb = load_workbook(USER_DATA_FILE)
    ws = wb.active
    ws.append(['Username', 'Password'])
    wb.save(USER_DATA_FILE)

if not os.path.exists(BILL_DATA_FILE):
    wb = load_workbook(BILL_DATA_FILE)
    ws = wb.active
    ws.append(['Patient ID', 'Admission Fees', 'Food Fees', 'Doctor Fees'])
    wb.save(BILL_DATA_FILE)


class LoginScreen(Screen):
    def __init__(self, **kwargs):
        super(LoginScreen, self).__init__(**kwargs)
        layout = BoxLayout(orientation='vertical')
        
        self.username = TextInput(hint_text='Username')
        self.password = TextInput(hint_text='Password', password=True)
        
        self.message = Label()
        
        login_btn = Button(text='Login')
        login_btn.bind(on_press=self.login)
        
        register_btn = Button(text='Register')
        register_btn.bind(on_press=self.register)
        
        layout.add_widget(self.username)
        layout.add_widget(self.password)
        layout.add_widget(login_btn)
        layout.add_widget(register_btn)
        layout.add_widget(self.message)
        
        self.add_widget(layout)
    
    def login(self, instance):
        username = self.username.text
        password = self.password.text
        
        wb = load_workbook(USER_DATA_FILE)
        ws = wb.active
        
        for row in ws.iter_rows(values_only=True):
            if row[0] == username and row[1] == password:
                self.manager.current = 'home'
                return
        self.message.text = 'Invalid Username or Password'
    
    def register(self, instance):
        username = self.username.text
        password = self.password.text
        
        wb = load_workbook(USER_DATA_FILE)
        ws = wb.active
        
        for row in ws.iter_rows(values_only=True):
            if row[0] == username:
                self.message.text = 'Username already exists'
                return
        
        ws.append([username, password])
        wb.save(USER_DATA_FILE)
        
        self.message.text = 'Registration Successful'


class HomeScreen(Screen):
    def __init__(self, **kwargs):
        super(HomeScreen, self).__init__(**kwargs)
        layout = BoxLayout(orientation='vertical')
        
        add_record_btn = Button(text='Add Record')
        add_record_btn.bind(on_press=self.go_to_add_record)
        
        schedule_appointment_btn = Button(text='Appointment Scheduling')
        schedule_appointment_btn.bind(on_press=self.go_to_schedule_appointment)
        
        bill_generation_btn = Button(text='Bill Generation')
        bill_generation_btn.bind(on_press=self.go_to_bill_generation)
        
        layout.add_widget(add_record_btn)
        layout.add_widget(schedule_appointment_btn)
        layout.add_widget(bill_generation_btn)
        
        self.add_widget(layout)
    
    def go_to_add_record(self, instance):
        self.manager.current = 'add_record'
    
    def go_to_schedule_appointment(self, instance):
        self.manager.current = 'schedule_appointment'
    
    def go_to_bill_generation(self, instance):
        self.manager.current = 'bill_generation'


class AddRecordScreen(Screen):
    def __init__(self, **kwargs):
        super(AddRecordScreen, self).__init__(**kwargs)
        layout = GridLayout(cols=2)
        
        self.patient_name = TextInput(hint_text='Patient Name')
        self.doctor_id = TextInput(hint_text='Doctor ID')
        self.treatment = TextInput(hint_text='Treatment Prescribed')
        
        save_btn = Button(text='Save Record')
        save_btn.bind(on_press=self.save_record)
        
        layout.add_widget(Label(text='Patient Name'))
        layout.add_widget(self.patient_name)
        layout.add_widget(Label(text='Doctor ID'))
        layout.add_widget(self.doctor_id)
        layout.add_widget(Label(text='Treatment Prescribed'))
        layout.add_widget(self.treatment)
        layout.add_widget(save_btn)
        
        self.add_widget(layout)
    
    def save_record(self, instance):
        patient_name = self.patient_name.text
        doctor_id = self.doctor_id.text
        treatment = self.treatment.text
        
        # Here, you should save the record to a file or database (omitted for simplicity)
        print(f"Record saved: {patient_name}, {doctor_id}, {treatment}")
        self.manager.current = 'home'


class ScheduleAppointmentScreen(Screen):
    def __init__(self, **kwargs):
        super(ScheduleAppointmentScreen, self).__init__(**kwargs)
        layout = BoxLayout(orientation='vertical')
        
        self.doctor_schedule = [
            {'name': 'Dr. Smith', 'dates': ['2024-08-01', '2024-08-02', '2024-08-03'], 'busy_dates': ['2024-08-01']},
            {'name': 'Dr. Johnson', 'dates': ['2024-08-01', '2024-08-02', '2024-08-03'], 'busy_dates': ['2024-08-02']},
            {'name': 'Dr. Lee', 'dates': ['2024-08-01', '2024-08-02', '2024-08-03'], 'busy_dates': []}
        ]
        
        for doctor in self.doctor_schedule:
            doc_label = Label(text=doctor['name'])
            layout.add_widget(doc_label)
            for date in doctor['dates']:
                color = (1, 0, 0, 1) if date in doctor['busy_dates'] else (0, 1, 0, 1)
                date_btn = Button(text=date, color=color)
                if date not in doctor['busy_dates']:
                    date_btn.bind(on_press=self.schedule_appointment)
                layout.add_widget(date_btn)
        
        back_btn = Button(text='Back')
        back_btn.bind(on_press=self.go_back)
        
        layout.add_widget(back_btn)
        
        self.add_widget(layout)
    
    def schedule_appointment(self, instance):
        selected_date = instance.text
        selected_doctor = None
        for doctor in self.doctor_schedule:
            if selected_date in doctor['dates'] and selected_date not in doctor['busy_dates']:
                selected_doctor = doctor['name']
                break
        
        if selected_doctor:
            receipt = f'Appointment scheduled with {selected_doctor} on {selected_date}.'
            self.manager.current = 'appointment_receipt'
            self.manager.get_screen('appointment_receipt').update_receipt(receipt)
    
    def go_back(self, instance):
        self.manager.current = 'home'


class AppointmentReceiptScreen(Screen):
    def __init__(self, **kwargs):
        super(AppointmentReceiptScreen, self).__init__(**kwargs)
        self.layout = BoxLayout(orientation='vertical')
        self.receipt_label = Label()
        back_btn = Button(text='Back to Home')
        back_btn.bind(on_press=self.go_back)
        
        self.layout.add_widget(self.receipt_label)
        self.layout.add_widget(back_btn)
        
        self.add_widget(self.layout)
    
    def update_receipt(self, receipt):
        self.receipt_label.text = receipt
    
    def go_back(self, instance):
        self.manager.current = 'home'


class BillGenerationScreen(Screen):
    def __init__(self, **kwargs):
        super(BillGenerationScreen, self).__init__(**kwargs)
        layout = GridLayout(cols=2)
        
        self.patient_id = TextInput(hint_text='Patient ID')
        self.total_bill = Label(text='Total Bill:')
        
        calculate_btn = Button(text='Calculate Bill')
        calculate_btn.bind(on_press=self.calculate_bill)
        
        layout.add_widget(Label(text='Patient ID'))
        layout.add_widget(self.patient_id)
        layout.add_widget(calculate_btn)
        layout.add_widget(self.total_bill)
        
        self.add_widget(layout)
    
    def calculate_bill(self, instance):
        patient_id = self.patient_id.text.strip()  # Remove any leading/trailing whitespace
        
        try:
            wb = load_workbook(BILL_DATA_FILE)
            ws = wb.active
            
            bill_found = False
            for row in ws.iter_rows(values_only=True):
                # Assuming patient ID in the Excel file is stored as a string
                if str(row[0]) == patient_id:
                    admission_fees = row[1]
                    food_fees = row[2]
                    doctor_fees = row[3]
                    total = admission_fees + food_fees + doctor_fees
                    self.total_bill.text = f'Total Bill: {total}'
                    bill_found = True
                    break
            
            if not bill_found:
                self.total_bill.text = 'Patient ID not found'
        except Exception as e:
            self.total_bill.text = f'Error: {str(e)}'


class MyApp(App):
    def build(self):
        sm = ScreenManager()
        sm.add_widget(LoginScreen(name='login'))
        sm.add_widget(HomeScreen(name='home'))
        sm.add_widget(AddRecordScreen(name='add_record'))
        sm.add_widget(ScheduleAppointmentScreen(name='schedule_appointment'))
        sm.add_widget(AppointmentReceiptScreen(name='appointment_receipt'))
        sm.add_widget(BillGenerationScreen(name='bill_generation'))
        
        return sm


if __name__ == '__main__':
    MyApp().run()
